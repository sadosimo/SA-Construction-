import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import sqlite3
import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

st.set_page_config(
    page_title="Sendas Antiguas", 
    page_icon="⛪", 
    layout="wide",
    menu_items={
        'About': 'Iglesia Sendas Antiguas'
    }
)

st.markdown("""
    <style>
    /* Modern glassmorphism theme */
    .stApp {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f0f23 100%) !important;
    }
    .stMetric {
        background: rgba(255,255,255,0.06) !important;
        padding: 12px 15px !important;
        border-radius: 12px !important;
        border: 1px solid rgba(255,255,255,0.1) !important;
        backdrop-filter: blur(10px);
    }
    .stMetric > div > div > div {
        color: white !important;
    }
    h1, h2, h3, h4, h5, h6 {
        color: #ffffff !important;
    }
    .stDataFrame {
        background: rgba(255,255,255,0.05);
        border-radius: 10px;
    }
    /* Glass cards */
    .glass-card {
        background: rgba(255,255,255,0.08) !important;
        backdrop-filter: blur(20px) !important;
        border-radius: 16px !important;
        border: 1px solid rgba(255,255,255,0.1) !important;
    }
    </style>
    """, unsafe_allow_html=True)

CARPETA_DATOS = "datos_iglesia"
CARPETA_IMAGENES = f"{CARPETA_DATOS}/facturas"
ARCHIVO_DB = f"{CARPETA_DATOS}/datos.db"

if not os.path.exists(CARPETA_DATOS):
    os.makedirs(CARPETA_DATOS)
if not os.path.exists(CARPETA_IMAGENES):
    os.makedirs(CARPETA_IMAGENES)

def init_db():
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS donaciones
                 (id INTEGER PRIMARY KEY, fecha TEXT, donante TEXT, monto REAL, metodo TEXT, notas TEXT, imagen TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS gastos
                 (id INTEGER PRIMARY KEY, fecha TEXT, descripcion TEXT, categoria TEXT, monto REAL, proveedor TEXT, imagen TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS etapas
                 (id INTEGER PRIMARY KEY, nombre TEXT, descripcion TEXT, presupuesto REAL, 
                  avance INTEGER DEFAULT 0, estado TEXT DEFAULT 'pendiente', fecha_inicio TEXT, fecha_fin TEXT,
                  fecha_fin_real TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS proyecto
                 (id INTEGER PRIMARY KEY, nombre TEXT, fecha_inicio TEXT, fecha_fin TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS tareas
                 (id INTEGER PRIMARY KEY, etapa_id INTEGER, nombre TEXT, completada INTEGER DEFAULT 0, 
                  FOREIGN KEY(etapa_id) REFERENCES etapas(id))''')
    
    # Agregar columnas si no existen (para bases de datos existentes)
    try:
        c.execute("ALTER TABLE donaciones ADD COLUMN imagen TEXT")
    except:
        pass
    try:
        c.execute("ALTER TABLE gastos ADD COLUMN imagen TEXT")
    except:
        pass
    
    conn.commit()
    conn.close()

def get_etapas():
    conn = sqlite3.connect(ARCHIVO_DB)
    df = pd.read_sql_query("SELECT * FROM etapas ORDER BY id", conn)
    conn.close()
    return df

def get_proyecto():
    conn = sqlite3.connect(ARCHIVO_DB)
    df = pd.read_sql_query("SELECT * FROM proyecto WHERE id = 1", conn)
    conn.close()
    return df

def actualizar_proyecto(nombre, fecha_inicio, fecha_fin):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO proyecto (id, nombre, fecha_inicio, fecha_fin) VALUES (1, ?, ?, ?)",
              (nombre, fecha_inicio, fecha_fin))
    conn.commit()
    conn.close()

def agregar_etapa(nombre, descripcion, presupuesto, fecha_inicio, fecha_fin):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("INSERT INTO etapas (nombre, descripcion, presupuesto, fecha_inicio, fecha_fin) VALUES (?, ?, ?, ?, ?)",
              (nombre, descripcion, presupuesto, fecha_inicio, fecha_fin))
    conn.commit()
    conn.close()

def actualizar_avance(etapa_id, avance):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    estado = 'completado' if avance == 100 else 'en_progreso' if avance > 0 else 'pendiente'
    c.execute("UPDATE etapas SET avance = ?, estado = ? WHERE id = ?", (avance, estado, etapa_id))
    conn.commit()
    conn.close()

def get_tareas(etapa_id):
    conn = sqlite3.connect(ARCHIVO_DB)
    df = pd.read_sql_query("SELECT * FROM tareas WHERE etapa_id = ? ORDER BY id", conn, params=(etapa_id,))
    conn.close()
    return df

def agregar_tarea(etapa_id, nombre):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("INSERT INTO tareas (etapa_id, nombre) VALUES (?, ?)", (etapa_id, nombre))
    conn.commit()
    conn.close()

def toggle_tarea(tarea_id):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("SELECT completada, etapa_id FROM tareas WHERE id = ?", (tarea_id,))
    result = c.fetchone()
    if result:
        new_status = 0 if result[0] == 1 else 1
        etapa_id = result[1]
        c.execute("UPDATE tareas SET completada = ? WHERE id = ?", (new_status, tarea_id))
        conn.commit()
        conn.close()
        actualizar_progreso_etapa(etapa_id)
    else:
        conn.close()

def eliminar_tarea(tarea_id):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("SELECT etapa_id FROM tareas WHERE id = ?", (tarea_id,))
    etapa_id_result = c.fetchone()
    c.execute("DELETE FROM tareas WHERE id = ?", (tarea_id,))
    conn.commit()
    conn.close()
    if etapa_id_result:
        actualizar_progreso_etapa(etapa_id_result[0])

def actualizar_progreso_etapa(etapa_id):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) as total, SUM(completada) as completa FROM tareas WHERE etapa_id = ?", (etapa_id,))
    result = c.fetchone()
    if result and result[0] > 0:
        total = result[0]
        completadas = result[1] if result[1] else 0
        avance = int((completadas / total) * 100)
    else:
        avance = 0
    estado = 'completado' if avance == 100 else 'en_progreso' if avance > 0 else 'pendiente'
    c.execute("UPDATE etapas SET avance = ?, estado = ? WHERE id = ?", (avance, estado, etapa_id))
    conn.commit()
    conn.close()
    return avance

def actualizar_donacion(id_val, fecha, donante, monto, metodo, notas):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("UPDATE donaciones SET fecha = ?, donante = ?, monto = ?, metodo = ?, notas = ? WHERE id = ?",
              (fecha, donante, monto, metodo, notas, id_val))
    conn.commit()
    conn.close()

def actualizar_gasto(id_val, fecha, descripcion, categoria, monto, proveedor):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("UPDATE gastos SET fecha = ?, descripcion = ?, categoria = ?, monto = ?, proveedor = ? WHERE id = ?",
              (fecha, descripcion, categoria, monto, proveedor, id_val))
    conn.commit()
    conn.close()

def actualizar_etapa(id_val, nombre, descripcion, presupuesto, fecha_inicio, fecha_fin):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("UPDATE etapas SET nombre = ?, descripcion = ?, presupuesto = ?, fecha_inicio = ?, fecha_fin = ? WHERE id = ?",
              (nombre, descripcion, presupuesto, fecha_inicio, fecha_fin, id_val))
    conn.commit()
    conn.close()

def eliminar_etapa(etapa_id):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("DELETE FROM etapas WHERE id = ?", (etapa_id,))
    conn.commit()
    conn.close()

def get_donaciones():
    conn = sqlite3.connect(ARCHIVO_DB)
    df = pd.read_sql_query("SELECT * FROM donaciones ORDER BY fecha DESC", conn)
    conn.close()
    return df

def get_gastos():
    conn = sqlite3.connect(ARCHIVO_DB)
    df = pd.read_sql_query("SELECT * FROM gastos ORDER BY fecha DESC", conn)
    conn.close()
    return df

def agregar_donacion(fecha, donante, monto, metodo, notas, imagen=None):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("INSERT INTO donaciones (fecha, donante, monto, metodo, notas, imagen) VALUES (?, ?, ?, ?, ?, ?)",
              (fecha, donante, monto, metodo, notas, imagen))
    conn.commit()
    conn.close()
    return c.lastrowid

def guardar_imagen(uploaded_file, prefijo):
    if uploaded_file is not None:
        try:
            ext = uploaded_file.name.split('.')[-1] if uploaded_file.name else 'png'
            nombre_archivo = f"{prefijo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
            ruta = os.path.join(CARPETA_IMAGENES, nombre_archivo)
            with open(ruta, "wb") as f:
                f.write(uploaded_file.getvalue())
            return nombre_archivo
        except Exception as e:
            print(f"Error guardando imagen: {e}")
            return None
    return None

def eliminar_imagen(nombre_archivo):
    if nombre_archivo and isinstance(nombre_archivo, str):
        try:
            ruta = os.path.join(CARPETA_IMAGENES, nombre_archivo)
            if os.path.exists(ruta):
                os.remove(ruta)
        except Exception as e:
            print(f"Error eliminando imagen: {e}")

def agregar_gasto(fecha, descripcion, categoria, monto, proveedor, imagen=None):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute("INSERT INTO gastos (fecha, descripcion, categoria, monto, proveedor, imagen) VALUES (?, ?, ?, ?, ?, ?)",
              (fecha, descripcion, categoria, monto, proveedor, imagen))
    conn.commit()
    conn.close()
    return c.lastrowid

def eliminar_registro(tabla, id_val):
    conn = sqlite3.connect(ARCHIVO_DB)
    c = conn.cursor()
    c.execute(f"DELETE FROM {tabla} WHERE id = ?", (id_val,))
    conn.commit()
    conn.close()

def generar_pdf(filtros=None):
    if filtros is None:
        filtros = {'donaciones': True, 'gastos': True, 'etapas': True, 'tareas': True, 'fecha_ini': '2024-01-01', 'fecha_fin': str(datetime.now().date())}
    
    fecha_ini = filtros.get('fecha_ini', '2024-01-01')
    fecha_fin = filtros.get('fecha_fin', str(datetime.now().date()))
    
    donativos = get_donaciones()
    gastos = get_gastos()
    etapas = get_etapas()
    
    # Filter by date
    if not donativos.empty:
        donativos = donativos[(donativos['fecha'] >= fecha_ini) & (donativos['fecha'] <= fecha_fin)]
    if not gastos.empty:
        gastos = gastos[(gastos['fecha'] >= fecha_ini) & (gastos['fecha'] <= fecha_fin)]
    
    total_donaciones = donativos['monto'].sum() if not donativos.empty else 0
    total_gastos = gastos['monto'].sum() if not gastos.empty else 0
    saldo = total_donaciones - total_gastos
    
    nombre_archivo = f"{CARPETA_DATOS}/reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    doc = SimpleDocTemplate(nombre_archivo, pagesize=letter)
    elementos = []
    estilos = getSampleStyleSheet()
    
    elementos.append(Paragraph("Reporte - Iglesia Sendas Antiguas", estilos['Title']))
    elementos.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", estilos['Normal']))
    elementos.append(Paragraph(f"Periodo: {fecha_ini} al {fecha_fin}", estilos['Normal']))
    elementos.append(Paragraph(" ", estilos['Normal']))
    
    elementos.append(Paragraph("RESUMEN FINANCIERO", estilos['Heading2']))
    datos_resumen = [
        ["Total Donaciones", f"${total_donaciones:,.2f}"],
        ["Total Gastos", f"${total_gastos:,.2f}"],
        ["SALDO", f"${saldo:,.2f}"]
    ]
    tabla_resumen = Table(datos_resumen, colWidths=[200, 150])
    tabla_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.green),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOLD', (0, -1), (-1, -1), True),
    ]))
    elementos.append(tabla_resumen)
    elementos.append(Paragraph(" ", estilos['Normal']))
    
    if filtros.get('donaciones', True) and not donativos.empty:
        elementos.append(Paragraph("DONACIONES", estilos['Heading2']))
        df_don = donativos[['fecha', 'donante', 'monto', 'metodo', 'notas']].copy()
        df_don.columns = ['Fecha', 'Donante', 'Monto', 'Metodo', 'Notas']
        data_don = [list(df_don.columns)] + df_don.values.tolist()
        tabla_don = Table(data_don)
        tabla_don.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elementos.append(tabla_don)
        elementos.append(Paragraph(" ", estilos['Normal']))
    
    if filtros.get('gastos', True) and not gastos.empty:
        elementos.append(Paragraph("GASTOS", estilos['Heading2']))
        df_gas = gastos[['fecha', 'descripcion', 'categoria', 'monto', 'proveedor']].copy()
        df_gas.columns = ['Fecha', 'Descripcion', 'Categoria', 'Monto', 'Proveedor']
        data_gas = [list(df_gas.columns)] + df_gas.values.tolist()
        tabla_gas = Table(data_gas)
        tabla_gas.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#922b21')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elementos.append(tabla_gas)
    
    if filtros.get('etapas', True) and not etapas.empty:
        elementos.append(Paragraph("ETAPAS DEL PROYECTO", estilos['Heading2']))
        df_eta = etapas[['nombre', 'presupuesto', 'avance', 'estado']].copy()
        df_eta.columns = ['Nombre', 'Presupuesto', 'Avance', 'Estado']
        data_eta = [list(df_eta.columns)] + df_eta.values.tolist()
        tabla_eta = Table(data_eta)
        tabla_eta.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3797')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elementos.append(tabla_eta)
    
    doc.build(elementos)
    return nombre_archivo

st.markdown("""
<style>
    /* Colores modernos */
    :root {
        --primary-color: #6366f1;
        --secondary-color: #8b5cf6;
        --accent-color: #06b6d4;
        --success-color: #10b981;
        --danger-color: #ef4444;
        --dark-bg: #0f172a;
        --card-bg: #1e293b;
    }
    
    /* Fondo principal mejorado */
    .stApp {
        background: linear-gradient(135deg, #0f172a 0%, #1e1b4b 50%, #0f172a 100%);
        background-attachment: fixed;
    }
    
    /* Glassmorphism - Efecto vidrio líquido */
    .glass {
        background: rgba(255, 255, 255, 0.05);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        border: 1px solid rgba(255, 255, 255, 0.1);
        border-radius: 20px;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
    }
    
    /* Glassmorphism fuerte */
    .glass-strong {
        background: rgba(255, 255, 255, 0.08);
        backdrop-filter: blur(30px);
        -webkit-backdrop-filter: blur(30px);
        border: 1px solid rgba(255, 255, 255, 0.15);
        border-radius: 24px;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.5);
    }
    
    /* Botones modernos con glass */
    .stButton>button {
        background: linear-gradient(135deg, rgba(99, 102, 241, 0.8) 0%, rgba(139, 92, 246, 0.8) 100%);
        color: white;
        border-radius: 12px;
        border: 1px solid rgba(255, 255, 255, 0.2);
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        transition: all 0.3s ease;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
    }
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 30px rgba(99, 102, 241, 0.5);
        background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
    }
    
    /* Título principal con glow */
    .titulo-principal {
        font-size: 2.5rem;
        font-weight: 800;
        background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 50%, #06b6d4 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        padding: 1rem;
        text-shadow: 0 0 40px rgba(99, 102, 241, 0.5);
    }
    
    /* Cards con glassmorphism */
    .stMetric {
        background: rgba(30, 41, 59, 0.6);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        border-radius: 20px;
        padding: 1.5rem;
        border: 1px solid rgba(99, 102, 241, 0.3);
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3);
    }
    
    /* Sidebar con glass */
    [data-testid="stSidebar"] {
        background: rgba(15, 23, 42, 0.7);
        backdrop-filter: blur(30px);
        -webkit-backdrop-filter: blur(30px);
        border-right: 1px solid rgba(255, 255, 255, 0.1);
    }
    
    /* Inputs con glass */
    .stTextInput>div>div>input, .stTextArea>div>div>textarea, .stSelectbox>div>div>div {
        border-radius: 12px;
        border: 2px solid rgba(99, 102, 241, 0.3);
        background: rgba(30, 41, 59, 0.5);
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        color: white;
    }
    .stTextInput>div>div>input:focus, .stTextArea>div>div>textarea:focus {
        border-color: #6366f1;
        box-shadow: 0 0 0 4px rgba(99, 102, 241, 0.3);
    }
    
    /* Expander con glass */
    .streamlit-expanderHeader {
        background: rgba(30, 41, 59, 0.5);
        backdrop-filter: blur(15px);
        -webkit-backdrop-filter: blur(15px);
        border-radius: 16px;
        border: 1px solid rgba(99, 102, 241, 0.2);
        color: white;
    }
    
    /* DataFrames con glass */
    [data-testid="stDataFrame"] {
        background: rgba(30, 41, 59, 0.4);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        border-radius: 16px;
        border: 1px solid rgba(99, 102, 241, 0.2);
    }
    
    /* Divider personalizado */
    .stDivider {
        border-color: rgba(99, 102, 241, 0.3);
    }
    
    /* Efecto de partículas flotantes */
    @keyframes float {
        0%, 100% { transform: translateY(0px); }
        50% { transform: translateY(-20px); }
    }
</style>
""", unsafe_allow_html=True)

init_db()

# Detectar modo de solo visualización desde URL
query_params = st.query_params
modo_solo_lectura = query_params.get("view") == "true"

if modo_solo_lectura:
    st.markdown("""
    <div style="background: rgba(99, 102, 241, 0.2); padding: 10px; border-radius: 10px; margin-bottom: 20px; text-align: center;">
        👁️ <b>Modo Solo Visualización</b> - Los datos son de solo lectura
    </div>
    """, unsafe_allow_html=True)

# Modern menu styling
st.markdown("""
    <style>
        div[data-testid="stSidebar"] {
            background: rgba(255, 255, 255, 0.08) !important;
            backdrop-filter: blur(20px) !important;
            border-right: 1px solid rgba(255, 255, 255, 0.15) !important;
        }
        [data-testid="stRadio"] > div {
            background: transparent;
        }
        [data-testid="stRadio"] label {
            background: rgba(255, 255, 255, 0.05) !important;
            padding: 10px 15px !important;
            margin: 3px 0 !important;
            border-radius: 8px !important;
            border: 1px solid rgba(255, 255, 255, 0.1) !important;
            transition: all 0.3s ease !important;
            font-size: 14px !important;
        }
        [data-testid="stRadio"] label:hover {
            background: rgba(99, 102, 241, 0.3) !important;
            border-color: rgba(99, 102, 241, 0.5) !important;
        }
    </style>
""", unsafe_allow_html=True)

# Elegant header - balanced size
st.sidebar.markdown("""
    <div style="
        background: linear-gradient(135deg, rgba(99, 102, 241, 0.4) 0%, rgba(139, 92, 246, 0.4) 100%);
        padding: 14px 12px;
        border-radius: 12px;
        text-align: center;
        margin-bottom: 15px;
        border: 1px solid rgba(255, 255, 255, 0.2);
        box-shadow: 0 4px 16px rgba(99, 102, 241, 0.3);
    ">
        <h2 style="color: white; margin: 0; font-size: 20px; font-weight: 800;">IGLESIA</h2>
        <p style="color: rgba(255,255,255,0.9); margin: 4px 0 0 0; font-size: 14px; font-weight: 600;">Sendas Antiguas</p>
        <p style="color: rgba(255,255,255,0.7); margin: 2px 0 0 0; font-size: 11px;">Control Construcción</p>
    </div>
""", unsafe_allow_html=True)

menu = st.sidebar.radio("", ["Dashboard Proyecto", "Finanzas", "Donaciones", "Gastos", "Proyecto", "PDF"], label_visibility="collapsed")

st.sidebar.markdown("---")
st.sidebar.markdown("""
    <div style="
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.2) 0%, rgba(5, 150, 105, 0.2) 100%);
        padding: 15px;
        border-radius: 12px;
        text-align: center;
        border: 1px solid rgba(16, 185, 129, 0.3);
    ">
        <p style="color: white; margin: 0; font-weight: 600;">Compartir Vista</p>
    </div>
""", unsafe_allow_html=True)
url_solo_lectura = f"{st.query_params.get('server_url', 'http://localhost:8501')}?view=true"
st.sidebar.code(url_solo_lectura, language=None)
st.sidebar.caption("Copia el enlace")

if menu == "Finanzas":
    st.header("Dashboard de Ingresos y Egresos")
    donativos = get_donaciones()
    gastos = get_gastos()
    
    total_donaciones = donativos['monto'].sum() if not donativos.empty else 0
    total_gastos = gastos['monto'].sum() if not gastos.empty else 0
    saldo = total_donaciones - total_gastos
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Donaciones", f"${total_donaciones:,.2f}", "💚")
    with col2:
        st.metric("Total Gastos", f"${total_gastos:,.2f}", "🔴", delta_color="inverse")
    with col3:
        st.metric("SALDO", f"${saldo:,.2f}", delta=f"${saldo:,.2f}")
    
    st.divider()
    
    c1, c2 = st.columns(2)
    
    with c1:
        st.subheader("📊 Ingresos vs Egresos")
        if not donativos.empty or not gastos.empty:
            df_grafico = pd.DataFrame({
                'Tipo': ['Ingresos', 'Egresos'],
                'Monto': [total_donaciones, total_gastos]
            })
            fig = px.pie(df_grafico, values='Monto', names='Tipo', 
                        color=['#10b981', '#ef4444'],
                        hole=0.4, title="📊 Distribución General")
            fig.update_traces(textinfo='percent+label', textfont=dict(size=14, color='white'))
            fig.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='white'),
                title=dict(font=dict(size=18, color='white'))
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay datos registrados aún")
    
    with c2:
        st.subheader("📈 Evolución Financiera")
        if not donativos.empty or not gastos.empty:
            df_linea = pd.DataFrame()
            if not donativos.empty:
                donativos['tipo'] = 'Ingreso'
                donativos['acumulado'] = donativos['monto'].cumsum()
                df_linea = donativos[['fecha', 'acumulado', 'tipo']].copy()
            if not gastos.empty:
                gastos['tipo'] = 'Gasto'
                gastos['acumulado_neg'] = -gastos['monto'].cumsum()
                if df_linea.empty:
                    df_linea = gastos[['fecha', 'acumulado_neg', 'tipo']].copy()
                else:
                    df_linea = pd.concat([df_linea, gastos[['fecha', 'acumulado_neg', 'tipo']]])
            
            if not df_linea.empty:
                df_linea = df_linea.sort_values('fecha')
                fig2 = px.line(df_linea, x='fecha', y='acumulado' if 'acumulado' in str(df_linea.columns) else 'acumulado_neg', 
                              markers=True, title="💹 Acumulado en el Tiempo")
                fig2.update_layout(
                    paper_bgcolor='rgba(0,0,0,0)',
                    plot_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='white'),
                    xaxis=dict(showgrid=True, gridcolor='rgba(255,255,255,0.1)', title_font=dict(color='white')),
                    yaxis=dict(showgrid=True, gridcolor='rgba(255,255,255,0.1)', title_font=dict(color='white')),
                    title=dict(font=dict(size=18, color='white'))
                )
                fig2.update_traces(line=dict(color='#6366f1', width=3))
                st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("No hay datos registrados aún")
    
    if not gastos.empty:
        st.divider()
        st.subheader("🏗️ Gastos por Categoría")
        cat_gastos = gastos.groupby('categoria')['monto'].sum().reset_index()
        fig3 = px.bar(cat_gastos, x='categoria', y='monto', 
                     color='monto', color_continuous_scale='Viridis',
                     title="📊 Gastos por Categoría")
        fig3.update_layout(
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(color='white'),
            xaxis=dict(showgrid=True, gridcolor='rgba(255,255,255,0.1)', title_font=dict(color='white')),
            yaxis=dict(showgrid=True, gridcolor='rgba(255,255,255,0.1)', title_font=dict(color='white')),
            title=dict(font=dict(size=18, color='white'))
        )
        st.plotly_chart(fig3, use_container_width=True)

if menu == "Donaciones":
    st.header("Registro de Donaciones")
    
    if not modo_solo_lectura:
        with st.form("form_donacion"):
            c1, c2 = st.columns(2)
            with c1:
                fecha = st.date_input("Fecha", datetime.now())
                donante = st.text_input("Nombre del Donante")
                imagen = st.file_uploader("Subir Factura/Comprobante", type=['png', 'jpg', 'jpeg', 'pdf'])
            with c2:
                monto = st.number_input("Monto ($)", min_value=0.0, step=100.0)
                metodo = st.selectbox("Metodo", ["Efectivo", "Transferencia", "Cheque", "Materiales", "Herramientas", "Trabajo", "Especie", "Otro"])
            notas = st.text_area("Notas (opcional)")
            enviado = st.form_submit_button("Registrar Donación")
            
            if enviado and donante and monto > 0:
                nombre_imagen = guardar_imagen(imagen, f"donacion_{donante.replace(' ', '_')}")
                agregar_donacion(fecha.strftime("%Y-%m-%d"), donante, monto, metodo, notas, nombre_imagen)
                st.success("✅ Donación registrada!")
                st.rerun()
    else:
        st.info("👁️ En modo solo visualización no puedes registrar nuevas donaciones")
    
    st.divider()
    st.subheader("📋 Historial de Donaciones")
    donativos = get_donaciones()
    if not donativos.empty:
        for i, row in donativos.iterrows():
            with st.expander(f"{row['fecha']} - {row['donante']} - ${row['monto']:,.2f}"):
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.write(f"**Metodo:** {row['metodo']}")
                    st.write(f"**Notas:** {row['notas'] if row['notas'] else 'Sin notas'}")
                    imagen_val = row.get('imagen')
                    if imagen_val and isinstance(imagen_val, str) and imagen_val.strip():
                        ruta_img = os.path.join(CARPETA_IMAGENES, imagen_val)
                        if os.path.exists(ruta_img):
                            st.image(ruta_img, caption="Factura/Comprobante", width=300)
                with col2:
                    col3, col4 = st.columns(2)
                    with col3:
                        if not modo_solo_lectura and st.button("✏️", key=f"edit_don_{row['id']}", help="Editar"):
                            st.session_state[f'editando_donacion_{row["id"]}'] = True
                    with col4:
                        if not modo_solo_lectura and st.button("🗑️", key=f"del_don_{row['id']}"):
                            if row.get('imagen'):
                                eliminar_imagen(row['imagen'])
                            eliminar_registro("donaciones", row['id'])
                            st.rerun()
                
                # Formulario de edición
                if st.session_state.get(f'editando_donacion_{row["id"]}', False):
                    st.divider()
                    with st.form(f"form_edit_don_{row['id']}"):
                        st.write("✏️ **Editar Donación**")
                        c1, c2 = st.columns(2)
                        with c1:
                            fecha_edit = st.date_input("Fecha", datetime.strptime(row['fecha'], '%Y-%m-%d'), key=f"fecha_don_{row['id']}")
                            donante_edit = st.text_input("Donante", row['donante'], key=f"donante_don_{row['id']}")
                        with c2:
                            monto_edit = st.number_input("Monto ($)", min_value=0.0, value=float(row['monto']), step=100.0, key=f"monto_don_{row['id']}")
                            metodo_edit = st.selectbox("Metodo", ["Efectivo", "Transferencia", "Cheque", "Materiales", "Herramientas", "Trabajo", "Especie", "Otro"], 
                                                       index=["Efectivo", "Transferencia", "Cheque", "Materiales", "Herramientas", "Trabajo", "Especie", "Otro"].index(row['metodo']) if row['metodo'] in ["Efectivo", "Transferencia", "Cheque", "Materiales", "Herramientas", "Trabajo", "Especie", "Otro"] else 0,
                                                       key=f"metodo_don_{row['id']}")
                        notas_edit = st.text_area("Notas", row['notas'] if row['notas'] else '', key=f"notas_don_{row['id']}")
                        
                        c3, c4 = st.columns(2)
                        with c3:
                            guardar = st.form_submit_button("💾 Guardar")
                        with c4:
                            cancelar = st.form_submit_button("❌ Cancelar")
                        
                        if guardar:
                            actualizar_donacion(row['id'], fecha_edit.strftime("%Y-%m-%d"), donante_edit, monto_edit, metodo_edit, notas_edit)
                            st.success("✅ Donación actualizada!")
                            st.session_state[f'editando_donacion_{row["id"]}'] = False
                            st.rerun()
                        if cancelar:
                            st.session_state[f'editando_donacion_{row["id"]}'] = False
                            st.rerun()
    else:
        st.info("No hay donaciones registradas")

if menu == "Gastos":
    st.header("Registro de Gastos")
    
    if not modo_solo_lectura:
        with st.form("form_gasto"):
            c1, c2 = st.columns(2)
            with c1:
                fecha = st.date_input("Fecha", datetime.now())
                descripcion = st.text_input("Descripción")
                imagen = st.file_uploader("📎 Subir Factura/Comprobante", type=['png', 'jpg', 'jpeg', 'pdf'])
            with c2:
                categoria = st.selectbox("Categoría", ["Materiales", "Mano de Obra", "Herramientas", "Transporte", "Permisos", "Otros"])
                monto = st.number_input("Monto ($)", min_value=0.0, step=100.0)
            proveedor = st.text_input("Proveedor (opcional)")
            enviado = st.form_submit_button("Registrar Gasto")
            
            if enviado and descripcion and monto > 0:
                nombre_imagen = guardar_imagen(imagen, f"gasto_{descripcion.replace(' ', '_')}")
                agregar_gasto(fecha.strftime("%Y-%m-%d"), descripcion, categoria, monto, proveedor, nombre_imagen)
                st.success("✅ Gasto registrado!")
                st.rerun()
    else:
        st.info("👁️ En modo solo visualización no puedes registrar nuevos gastos")
    
    st.divider()
    st.subheader("📋 Historial de Gastos")
    gastos = get_gastos()
    if not gastos.empty:
        for i, row in gastos.iterrows():
            with st.expander(f"{row['fecha']} - {row['descripcion']} - ${row['monto']:,.2f}"):
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.write(f"**Categoria:** {row['categoria']}")
                    st.write(f"**Proveedor:** {row['proveedor'] if row['proveedor'] else 'No especificado'}")
                    imagen_gas = row.get('imagen')
                    if imagen_gas and isinstance(imagen_gas, str) and imagen_gas.strip():
                        ruta_img = os.path.join(CARPETA_IMAGENES, imagen_gas)
                        if os.path.exists(ruta_img):
                            st.image(ruta_img, caption="Factura/Comprobante", width=300)
                with col2:
                    col3, col4 = st.columns(2)
                    with col3:
                        if not modo_solo_lectura and st.button("✏️", key=f"edit_gas_{row['id']}", help="Editar"):
                            st.session_state[f'editando_gasto_{row["id"]}'] = True
                    with col4:
                        if not modo_solo_lectura and st.button("🗑️", key=f"del_gas_{row['id']}"):
                            if row.get('imagen'):
                                eliminar_imagen(row['imagen'])
                            eliminar_registro("gastos", row['id'])
                            st.rerun()
                
                # Formulario de edición
                if st.session_state.get(f'editando_gasto_{row["id"]}', False):
                    st.divider()
                    with st.form(f"form_edit_gas_{row['id']}"):
                        st.write("✏️ **Editar Gasto**")
                        c1, c2 = st.columns(2)
                        categorias = ["Materiales", "Mano de Obra", "Herramientas", "Transporte", "Permisos", "Otros"]
                        with c1:
                            fecha_edit = st.date_input("Fecha", datetime.strptime(row['fecha'], '%Y-%m-%d'), key=f"fecha_gas_{row['id']}")
                            descripcion_edit = st.text_input("Descripción", row['descripcion'], key=f"desc_gas_{row['id']}")
                            categoria_edit = st.selectbox("Categoría", categorias, 
                                                          index=categorias.index(row['categoria']) if row['categoria'] in categorias else 0,
                                                          key=f"cat_gas_{row['id']}")
                        with c2:
                            monto_edit = st.number_input("Monto ($)", min_value=0.0, value=float(row['monto']), step=100.0, key=f"monto_gas_{row['id']}")
                            proveedor_edit = st.text_input("Proveedor", row['proveedor'] if row['proveedor'] else '', key=f"prov_gas_{row['id']}")
                        
                        c3, c4 = st.columns(2)
                        with c3:
                            guardar = st.form_submit_button("💾 Guardar")
                        with c4:
                            cancelar = st.form_submit_button("❌ Cancelar")
                        
                        if guardar:
                            actualizar_gasto(row['id'], fecha_edit.strftime("%Y-%m-%d"), descripcion_edit, categoria_edit, monto_edit, proveedor_edit)
                            st.success("✅ Gasto actualizado!")
                            st.session_state[f'editando_gasto_{row["id"]}'] = False
                            st.rerun()
                        if cancelar:
                            st.session_state[f'editando_gasto_{row["id"]}'] = False
                            st.rerun()
    else:
        st.info("No hay gastos registrados")

if menu == "Proyecto":
    st.header("Seguimiento del Proyecto")
    st.write("Administra las etapas de construcción de la iglesia")
    
    # Configuración del proyecto general
    proyecto = get_proyecto()
    
    if not modo_solo_lectura:
        with st.expander("📅 Fechas del Proyecto"):
            with st.form("form_proyecto"):
                st.write("**Fechas Límite del Proyecto**")
                c_p1, c_p2 = st.columns(2)
                with c_p1:
                    nombre_proyecto = st.text_input("Nombre del Proyecto", value=proyecto['nombre'].iloc[0] if not proyecto.empty else "Construcción Iglesia Sendas Antiguas")
                    fecha_inicio_proy = st.date_input("Fecha de Inicio", 
                        datetime.strptime(proyecto['fecha_inicio'].iloc[0], '%Y-%m-%d') if not proyecto.empty and proyecto['fecha_inicio'].iloc[0] else datetime.now())
                with c_p2:
                    fecha_fin_proy = st.date_input("Fecha de Fin Estimada",
                        datetime.strptime(proyecto['fecha_fin'].iloc[0], '%Y-%m-%d') if not proyecto.empty and proyecto['fecha_fin'].iloc[0] else datetime.now())
                
                if st.form_submit_button("💾 Guardar Fechas"):
                    actualizar_proyecto(nombre_proyecto, str(fecha_inicio_proy), str(fecha_fin_proy))
                    st.success("✅ Fechas guardadas")
                    st.rerun()
    else:
        if not proyecto.empty:
            col_d1, col_d2 = st.columns(2)
            with col_d1:
                st.info(f"📅 **Inicio:** {proyecto['fecha_inicio'].iloc[0]}")
            with col_d2:
                st.warning(f"⏰ **Fecha Límite:** {proyecto['fecha_fin'].iloc[0]}")
    
    st.divider()
    
    # Agregar nueva etapa
    if not modo_solo_lectura:
        with st.expander("➕ Agregar Nueva Etapa"):
            with st.form("form_etapa"):
                c1, c2 = st.columns(2)
                with c1:
                    nombre_etapa = st.text_input("Nombre de la Etapa")
                    presupuesto = st.number_input("Presupuesto ($)", min_value=0.0, step=100.0)
                with c2:
                    descripcion = st.text_area("Descripción")
                    fecha_inicio = st.date_input("Fecha de Inicio", datetime.now())
                
                fecha_fin = st.date_input("Fecha de Fin Estimada", datetime.now())
                
                if st.form_submit_button("➕ Agregar Etapa"):
                    if nombre_etapa and presupuesto > 0:
                        agregar_etapa(nombre_etapa, descripcion, presupuesto, str(fecha_inicio), str(fecha_fin))
                        st.success("✅ Etapa agregada correctamente")
                        st.rerun()
                    else:
                        st.error("⚠️ Ingresa el nombre y presupuesto de la etapa")
    else:
        st.info("👁️ En modo solo visualización no puedes agregar nuevas etapas")
    
    st.divider()
    
    # Mostrar etapas
    etapas = get_etapas()
    proyecto = get_proyecto()
    
    if not etapas.empty:
        # Resumen de progreso
        total_presupuesto = etapas['presupuesto'].sum()
        promedio_avance = etapas['avance'].mean()
        etapas_completadas = len(etapas[etapas['estado'] == 'completado'])
        etapas_en_progreso = len(etapas[etapas['estado'] == 'en_progreso'])
        
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.metric("Total Etapas", len(etapas))
        with c2:
            st.metric("Presupuesto Total", f"${total_presupuesto:,.2f}")
        with c3:
            st.metric("Completadas", etapas_completadas, f"{etapas_completadas}/{len(etapas)}")
        with c4:
            st.metric("Avance Promedio", f"{promedio_avance:.1f}%")
        with c5:
            if proyecto.empty or not proyecto['fecha_fin'].iloc[0]:
                st.metric("⏰ Días Restantes", "Sin fecha")
            else:
                fecha_fin = datetime.strptime(proyecto['fecha_fin'].iloc[0], '%Y-%m-%d')
                dias_restantes = (fecha_fin - datetime.now()).days
                if dias_restantes > 0:
                    st.metric("⏰ Días Restantes", f"{dias_restantes} días", delta_color="normal")
                elif dias_restantes == 0:
                    st.metric("⏰ Días Restantes", "¡Vence hoy!", delta_color="inverse")
                else:
                    st.metric("⏰ Días Restantes", f"-{abs(dias_restantes)}", delta_color="inverse")
        
        st.divider()
        
        # Barra de progreso general
        st.subheader("📊 Progreso General del Proyecto")
        st.progress(int(promedio_avance), text=f"Avance total: {promedio_avance:.1f}%")
        
        # Visual countdown timer
        if not proyecto.empty and proyecto['fecha_fin'].iloc[0]:
            fecha_fin = datetime.strptime(proyecto['fecha_fin'].iloc[0], '%Y-%m-%d')
            dias_restantes = (fecha_fin - datetime.now()).days
            
            # Timer visual
            col_timer1, col_timer2 = st.columns([1, 2])
            with col_timer1:
                if dias_restantes > 0:
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 20px; border-radius: 15px; text-align: center;">
                        <h1 style="color: white; margin: 0; font-size: 48px;">{dias_restantes}</h1>
                        <p style="color: white; margin: 0;">días restantes</p>
                    </div>
                    """, unsafe_allow_html=True)
                elif dias_restantes == 0:
                    st.markdown("""
                    <div style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); padding: 20px; border-radius: 15px; text-align: center;">
                        <h1 style="color: white; margin: 0; font-size: 48px;">0</h1>
                        <p style="color: white; margin: 0;">¡VENCE HOY!</p>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #ff0844 0%, #ffb199 100%); padding: 20px; border-radius: 15px; text-align: center;">
                        <h1 style="color: white; margin: 0; font-size: 48px;">{abs(dias_restantes)}</h1>
                        <p style="color: white; margin: 0;">días de retraso</p>
                    </div>
                    """, unsafe_allow_html=True)
        
        st.divider()
        
        # Listado de etapas
        st.subheader("📋 Etapas del Proyecto")
        
        for i, row in etapas.iterrows():
            # Color según estado
            if row['estado'] == 'completado':
                color_emoji = "✅"
                color_bar = "#10b981"
            elif row['estado'] == 'en_progreso':
                color_emoji = "🔄"
                color_bar = "#6366f1"
            else:
                color_emoji = "⏳"
                color_bar = "#64748b"
            
            with st.expander(f"{color_emoji} {row['nombre']} - {row['avance']}%"):
                c1, c2 = st.columns([3, 1])
                with c1:
                    st.write(f"**Descripción:** {row['descripcion'] if row['descripcion'] else 'Sin descripción'}")
                    st.write(f"**Presupuesto:** ${row['presupuesto']:,.2f}")
                    st.write(f"**Fecha inicio:** {row['fecha_inicio']}")
                    st.write(f"**Fecha fin estimada:** {row['fecha_fin']}")
                    
                    # Mostrar días restantes
                    if row['fecha_fin']:
                        try:
                            fecha_fin_etapa = datetime.strptime(row['fecha_fin'], '%Y-%m-%d')
                            dias_etapa = (fecha_fin_etapa - datetime.now()).days
                            if dias_etapa > 0:
                                st.warning(f"⏰ **Días restantes:** {dias_etapa} días")
                            elif dias_etapa == 0:
                                st.warning(f"⚠️ **¡Vence hoy!**")
                            else:
                                st.error(f"❌ **Vencido hace {abs(dias_etapa)} días**")
                        except:
                            pass
                    
                    st.progress(row['avance'] / 100, text=f"Avance: {row['avance']}%")
                with c2:
                    col3, col4 = st.columns(2)
                    with col3:
                        if not modo_solo_lectura and st.button("✏️", key=f"edit_etapa_{row['id']}", help="Editar"):
                            st.session_state[f'editando_etapa_{row["id"]}'] = True
                            st.rerun()
                    with col4:
                        if not modo_solo_lectura and st.button("🗑️", key=f"del_etapa_{row['id']}"):
                            eliminar_etapa(row['id'])
                            st.rerun()
                
                # Mostrar progreso basado en tareas
                tareas_etapa = get_tareas(row['id'])
                if tareas_etapa.empty:
                    if modo_solo_lectura:
                        st.write(f"**Avance:** {row['avance']}%")
                    else:
                        nuevo_avance = st.slider(f"Avance %", 0, 100, row['avance'], key=f"avance_{row['id']}")
                        if nuevo_avance != row['avance']:
                            actualizar_avance(row['id'], nuevo_avance)
                            st.rerun()
                else:
                    st.write(f"**Avance automático (checklist):** {row['avance']}%")
                
                # Checklist de tareas
                st.divider()
                st.subheader("☑️ Checklist de Tareas")
                
                if not modo_solo_lectura:
                    c_at1, c_at2 = st.columns([4, 1])
                    with c_at1:
                        nueva_tarea = st.text_input(f"Nueva tarea para: {row['nombre']}", key=f"input_tarea_{row['id']}", placeholder="Escribe una tarea...")
                    with c_at2:
                        if st.button("➕ Agregar", key=f"btn_agregar_tarea_{row['id']}"):
                            if nueva_tarea:
                                agregar_tarea(row['id'], nueva_tarea)
                                actualizar_progreso_etapa(row['id'])
                                st.rerun()
                
                if not tareas_etapa.empty:
                    total_tareas = len(tareas_etapa)
                    completadas = len(tareas_etapa[tareas_etapa['completada'] == 1])
                    
                    st.progress(completadas / total_tareas if total_tareas > 0 else 0, text=f"Tareas: {completadas}/{total_tareas} completadas")
                    
                    for _, tarea in tareas_etapa.iterrows():
                        c_chk1, c_chk2, c_chk3 = st.columns([1, 4, 1])
                        emoji = "✅" if tarea['completada'] == 1 else "⬜"
                        with c_chk1:
                            if not modo_solo_lectura:
                                if st.button(emoji, key=f"btn_toggle_{tarea['id']}"):
                                    toggle_tarea(tarea['id'])
                                    nuevo_avance = actualizar_progreso_etapa(row['id'])
                                    st.success(f"✅ Tarea actualizada! Avance: {nuevo_avance}%")
                                    st.rerun()
                            else:
                                st.write(emoji)
                        with c_chk2:
                            estado_texto = "~~" + tarea['nombre'] + "~~" if tarea['completada'] == 1 else tarea['nombre']
                            st.markdown(estado_texto)
                        with c_chk3:
                            if not modo_solo_lectura and st.button("🗑️", key=f"del_tarea_{tarea['id']}"):
                                nuevo_avance = actualizar_progreso_etapa(row['id'])
                                st.success(f"Tarea eliminada. Avance: {nuevo_avance}%")
                                st.rerun()
                else:
                    st.info("No hay tareas agregadas. ¡Agrega tareas para seguimiento!")
                
                st.divider()
                
                # Formulario de edición
                if st.session_state.get(f'editando_etapa_{row["id"]}', False):
                    st.divider()
                    with st.form(f"form_edit_etapa_{row['id']}"):
                        st.write("✏️ **Editar Etapa**")
                        c1, c2 = st.columns(2)
                        with c1:
                            nombre_edit = st.text_input("Nombre", row['nombre'], key=f"nom_etapa_{row['id']}")
                            presupuesto_edit = st.number_input("Presupuesto ($)", min_value=0.0, value=float(row['presupuesto']), step=100.0, key=f"pres_etapa_{row['id']}")
                        with c2:
                            descripcion_edit = st.text_area("Descripción", row['descripcion'] if row['descripcion'] else '', key=f"desc_etapa_{row['id']}")
                        
                        c3, c4 = st.columns(2)
                        with c3:
                            fecha_inicio_edit = st.date_input("Fecha de Inicio", datetime.strptime(row['fecha_inicio'], '%Y-%m-%d') if row['fecha_inicio'] else datetime.now(), key=f"ini_etapa_{row['id']}")
                        with c4:
                            fecha_fin_edit = st.date_input("Fecha de Fin", datetime.strptime(row['fecha_fin'], '%Y-%m-%d') if row['fecha_fin'] else datetime.now(), key=f"fin_etapa_{row['id']}")
                        
                        c5, c6 = st.columns(2)
                        with c5:
                            guardar = st.form_submit_button("💾 Guardar")
                        with c6:
                            cancelar = st.form_submit_button("❌ Cancelar")
                        
                        if guardar:
                            actualizar_etapa(row['id'], nombre_edit, descripcion_edit, presupuesto_edit, 
                                          str(fecha_inicio_edit), str(fecha_fin_edit))
                            st.success("✅ Etapa actualizada!")
                            st.session_state[f'editando_etapa_{row["id"]}'] = False
                            st.rerun()
                        if cancelar:
                            st.session_state[f'editando_etapa_{row["id"]}'] = False
                            st.rerun()
        
        # Gráfico de progreso por etapa
        st.divider()
        st.subheader("📈 Distribución del Presupuesto por Etapa")
        
        if not etapas.empty:
            fig_etapas = px.bar(etapas, x='nombre', y='presupuesto', 
                               color='avance', color_continuous_scale='Viridis',
                               title="Presupuesto por Etapa")
            fig_etapas.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='white'),
                xaxis=dict(showgrid=True, gridcolor='rgba(255,255,255,0.1)', title_font=dict(color='white')),
                yaxis=dict(showgrid=True, gridcolor='rgba(255,255,255,0.1)', title_font=dict(color='white'))
            )
            st.plotly_chart(fig_etapas, use_container_width=True)
    else:
        st.info("No hay etapas registradas. Agrega la primera etapa del proyecto.")

if menu == "Dashboard Proyecto":
    st.header("Dashboard del Proyecto")
    st.write("Resumen financiero y de progreso del proyecto de construcción")
    
    # Budget overview
    etapas = get_etapas()
    donativos = get_donaciones()
    gastos = get_gastos()
    proyecto = get_proyecto()
    
    total_presupuesto = etapas['presupuesto'].sum() if not etapas.empty else 0
    total_donaciones = donativos['monto'].sum() if not donativos.empty else 0
    total_gastos = gastos['monto'].sum() if not gastos.empty else 0
    saldo_disponible = total_donaciones - total_gastos
    faltante = total_presupuesto - saldo_disponible if total_presupuesto > 0 else 0
    
    promedio_avance = etapas['avance'].mean() if not etapas.empty else 0
    etapas_completadas = len(etapas[etapas['estado'] == 'completado']) if not etapas.empty else 0
    
    # Metrics
    dbc1, dbc2, dbc3, dbc4 = st.columns(4)
    with dbc1:
        st.metric("Presupuesto Total", f"${total_presupuesto:,.2f}")
    with dbc2:
        st.metric("Donaciones", f"${total_donaciones:,.2f}", delta=f"+${total_donaciones:,.2f}")
    with dbc3:
        st.metric("Gastos", f"${total_gastos:,.2f}", delta=f"-${total_gastos:,.2f}", delta_color="inverse")
    with dbc4:
        if faltante > 0:
            st.metric("Faltante", f"${faltante:,.2f}", delta_color="inverse")
        else:
            st.metric("Sobrante", f"${abs(faltante):,.2f}")
    
    st.divider()
    
    # Progress metrics
    pc1, pc2, pc3 = st.columns(3)
    with pc1:
        st.metric("Etapas Totales", len(etapas) if not etapas.empty else 0)
    with pc2:
        st.metric("Etapas Completadas", etapas_completadas)
    with pc3:
        st.metric("Avance del Proyecto", f"{promedio_avance:.1f}%")
    
    # Progress bar
    st.progress(promedio_avance / 100, text=f"Avance General: {promedio_avance:.1f}%")
    
    st.divider()
    
    # Deadline
    if not proyecto.empty and proyecto['fecha_fin'].iloc[0]:
        fecha_fin = datetime.strptime(proyecto['fecha_fin'].iloc[0], '%Y-%m-%d')
        dias_restantes = (fecha_fin - datetime.now()).days
        
        dc1, dc2 = st.columns(2)
        with dc1:
            if dias_restantes > 0:
                st.metric("Días Restantes", f"{dias_restantes} días")
            elif dias_restantes == 0:
                st.metric("Días Restantes", "¡Vence hoy!", delta_color="inverse")
            else:
                st.metric("Días de Retraso", f"{abs(dias_restantes)} días", delta_color="inverse")
        with dc2:
            st.metric("Fecha Límite", str(proyecto['fecha_fin'].iloc[0]))
    
    st.divider()
    
    # Budget progress
    if total_presupuesto > 0 and saldo_disponible > 0:
        presupuesto_progress = min((saldo_disponible / total_presupuesto) * 100, 100)
        st.progress(presupuesto_progress / 100, text=f"Presupuesto Cubierto: {presupuesto_progress:.1f}%")
    elif total_presupuesto > 0:
        st.progress(0, text="Presupuesto Cubierto: 0%")

elif menu == "PDF":
    st.header("Generar Reporte PDF")
    
    st.write("Selecciona que desea incluir en el reporte:")
    
    # Filters checkboxes
    c_f1, c_f2 = st.columns(2)
    with c_f1:
        incluir_donaciones = st.checkbox("Donaciones", value=True)
        incluir_etapas = st.checkbox("Etapas del Proyecto", value=True)
    with c_f2:
        incluir_gastos = st.checkbox("Gastos", value=True)
        incluir_tareas = st.checkbox("Tareas/Checklist", value=True)
    
    # Date range filter
    st.write(" Filtrar por fecha:")
    fecha_ini = st.date_input("Desde", value=datetime(2024,1,1))
    fecha_fin = st.date_input("Hasta", value=datetime.now())
    
    st.divider()
    
    # Preview counts
    donativos = get_donaciones()
    gastos = get_gastos()
    etapas = get_etapas()
    
    if not donativos.empty:
        donativos_fecha = donativos[(donativos['fecha'] >= str(fecha_ini)) & (donativos['fecha'] <= str(fecha_fin))]
    else:
        donativos_fecha = donativos
    
    if not gastos.empty:
        gastos_fecha = gastos[(gastos['fecha'] >= str(fecha_ini)) & (gastos['fecha'] <= str(fecha_fin))]
    else:
        gastos_fecha = gastos
    
    total_donaciones = donativos_fecha['monto'].sum() if not donativos_fecha.empty else 0
    total_gastos = gastos_fecha['monto'].sum() if not gastos_fecha.empty else 0
    saldo = total_donaciones - total_gastos
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Donaciones", f"${total_donaciones:,.2f}")
    with col2:
        st.metric("Gastos", f"${total_gastos:,.2f}")
    with col3:
        st.metric("SALDO", f"${saldo:,.2f}")
    
    # Export options for dashboards
    st.divider()
    st.write("Exportar Datos:")
    
    ec1, ec2, ec3, ec4 = st.columns(4)
    with ec1:
        if st.button("Donaciones CSV"):
            donativos = get_donaciones()
            if not donativos.empty:
                csv = donativos.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="CSV",
                    data=csv,
                    file_name=f"donaciones_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
    with ec2:
        if st.button("Gastos CSV"):
            gastos = get_gastos()
            if not gastos.empty:
                csv = gastos.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="CSV",
                    data=csv,
                    file_name=f"gastos_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
    with ec3:
        if st.button("Etapas CSV"):
            etapas = get_etapas()
            if not etapas.empty:
                csv = etapas.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="CSV",
                    data=csv,
                    file_name=f"etapas_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
    with ec4:
        if st.button("Excel Completo"):
            try:
                from openpyxl import Workbook
                from io import BytesIO
                wb = Workbook()
                
                ws1 = wb.active
                ws1.title = "Donaciones"
                donativos = get_donaciones()
                if donativos is not None and not donativos.empty:
                    ws1.append(['Fecha', 'Donante', 'Monto', 'Metodo', 'Notas'])
                    for _, row in donativos.iterrows():
                        ws1.append([row['fecha'], row['donante'], row['monto'], row['metodo'], str(row.get('notas', ''))])
                
                ws2 = wb.create_sheet("Gastos")
                gastos = get_gastos()
                if gastos is not None and not gastos.empty:
                    ws2.append(['Fecha', 'Descripcion', 'Categoria', 'Monto', 'Proveedor'])
                    for _, row in gastos.iterrows():
                        ws2.append([row['fecha'], row['descripcion'], row['categoria'], row['monto'], str(row.get('proveedor', ''))])
                
                ws3 = wb.create_sheet("Etapas")
                etapas = get_etapas()
                if etapas is not None and not etapas.empty:
                    ws3.append(['Nombre', 'Presupuesto', 'Avance', 'Estado', 'Fecha Inicio', 'Fecha Fin'])
                    for _, row in etapas.iterrows():
                        ws3.append([row['nombre'], row['presupuesto'], row['avance'], row['estado'], str(row.get('fecha_inicio', '')), str(row.get('fecha_fin', ''))])
                
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                st.download_button(
                    label="Excel",
                    data=buffer.getvalue(),
                    file_name=f"reporteCompleto_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openpyxl-formula"
                )
            except Exception as e:
                st.error(f"Excel no disponible: {str(e)}")
    
    st.divider()
    
    if modo_solo_lectura:
        st.info("En modo solo visualizacion no puedes generar PDFs")
    elif st.button("GENERAR PDF"):
        filtros = {
            'donaciones': incluir_donaciones,
            'gastos': incluir_gastos,
            'etapas': incluir_etapas,
            'tareas': incluir_tareas,
            'fecha_ini': str(fecha_ini),
            'fecha_fin': str(fecha_fin)
        }
        archivo = generar_pdf(filtros)
        st.success(f"Reporte generado: {archivo}")
        with open(archivo, "rb") as pdf:
            st.download_button(
                label="Descargar PDF",
                data=pdf,
                file_name=f"reporte_iglesia_{datetime.now().strftime('%Y%m%d')}.pdf",
                mime="application/pdf"
            )

st.markdown("---")
st.caption("Iglesia Sendas Antiguas | Para compartir en red local: streamlit run app.py --server.address 0.0.0.0")